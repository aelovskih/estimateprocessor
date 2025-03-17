import streamlit as st
import pandas as pd
import random
from io import BytesIO
import openpyxl
import math

#############################
# 1. Словарь почасовых ставок для грейдов
#############################
hourly_rates = {
    "Продуктовая аналитика стажер": 150,
    "UX-аналитик junior": 1250,
    "UX-аналитик middle -": 1650,
    "UX-аналитик middle": 2100,
    "UX-аналитик senior": 3050,
    "Web/mobile-аналитик junior": 1250,
    "Web/mobile-аналитик middle -": 1650,
    "Web/mobile-аналитик middle": 2100,
    "Web/mobile-аналитик middle +": 2650,
    "Web/mobile-аналитик senior": 3350,
    "Web/mobile-аналитик TL": 3600,
    "BI/data-аналитик junior": 1650,
    "BI/data-аналитик middle": 2500,
    "BI/data-аналитик middle +": 2900,
    "BI/data-аналитик senior": 3450,
    "Product manager junior": 2100,
    "Product manager middle": 2800,
    "Product manager middle +": 3200,
    "Product manager Senior": 3900,
    "Web/mobile-analyst middle": 3000,
    "Web/mobile-analyst senior": 3500,
    "Дизайнер стажер": 150,
    "Дизайнер junior-": 1150,
    "Дизайнер junior": 1400,
    "Дизайнер junior+": 1850,
    "Дизайнер middle-": 2150,
    "Дизайнер middle": 2550,
    "Дизайнер middle+": 3150,
    "Дизайнер senior": 3550,
    "Дизайнер senior+": 3850,
    "Art Director": 4250,
    "Web middle": 2300,
    "Web senior": 3000,
    "Системный аналитик Стажер": 150,
    "Системный аналитик junior-": 1300,
    "Системный аналитик junior": 1850,
    "Системный аналитик junior+": 2150,
    "Системный аналитик middle-": 2550,
    "Системный аналитик middle": 3150,
    "Системный аналитик middle+": 3400,
    "Системный аналитик senior": 3700,
    "Системный аналитик senior+": 4000,
    "Системный аналитик lead": 4400,
    "Проектировщик стажер": 150,
    "Проектировщик junior": 1850,
    "Проектировщик middle": 2300,
    "Проектировщик middle+": 2850,
    "Проектировщик senior": 3150,
    "Проектировщик senior+": 3550,
    "Проектировщик lead": 4000,
    "System analyst middle": 2750,
    "System analyst senior": 3200,
    "SEO": 3465,
    "tech.writer": 2805,
    "UX writer": 2805,
    "PM стажер": 150,
    "PM intern": 150,
    "PM junior": 900,
    "PM junior+": 1350,
    "PM middle": 1750,
    "PM middle+": 2000,
    "PM senior": 2450,
    "PM senior+": 2750,
    "GH": 3200,
    "PMO": 3550,
    "Bitrix junior": 850,
    "Bitrix junior+": 1100,
    "Bitrix middle-": 1450,
    "Bitrix middle": 1950,  # При повторении последнее значение берется
    "Bitrix middle+": 2400,
    "Bitrix senior-": 2700,
    "Bitrix senior": 2900,
    "Bitrix senior+": 3150,
    "Bitrix teamlead -": 2450,
    "Bitrix teamlead": 2850,
    "Bitrix teamlead +": 3150,
    "Bitrix teamlead grouphead -": 3450,
    "Bitrix teamlead grouphead": 3750,
    "Bitrix teamlead grouphead +": 4300,
    "Разработка стажер": 150,
    "Framework junior": 850,
    "Framework junior+": 1400,
    "Framework middle-": 1950,
    "Framework middle": 2500,
    "Framework middle+": 2900,
    "Framework senior-": 3300,
    "Framework senior": 3700,
    "Framework senior+": 4150,
    "Framework teamlead -": 3150,
    "Framework teamlead": 3450,
    "Framework teamlead +": 3750,
    "Framework teamlead grouphead -": 4000,
    "Framework teamlead grouphead": 4300,
    "Framework teamlead grouphead +": 4850,
    "QA junior-": 700,
    "QA junior": 950,
    "QA junior +": 1250,
    "QA middle-": 1500,
    "QA middle": 1700,
    "QA middle+": 1950,
    "QA senior -": 2150,
    "QA senior": 2350,
    "QA Teamlead": 2850,
    "AQA junior-": 950,
    "AQA junior": 1250,
    "AQA junior +": 1500,
    "AQA middle-": 1850,
    "AQA middle": 2200,
    "AQA middle+": 2550,
    "AQA senior -": 2900,
    "AQA senior": 3150,
    "AQA teamlead": 3600,
    "QA middle": 2400,
    "QA senior": 3200,
    "QA AT Middle": 3200,
    "QA AT Senior": 4000,
    "Front-end junior": 950,
    "Front-end junior +": 1250,
    "Front-end middle -": 1650,
    "Front-end middle": 2050,
    "Front-end middle +": 2500,
    "Front-end senior -": 2900,
    "Front-end senior": 3300,
    "Front-end senior +": 3700,
    "Front-end teamlead": 4300,
    "NodeJS junior": 950,
    "NodeJS junior +": 1250,
    "NodeJS middle -": 1650,
    "NodeJS middle": 2050,
    "NodeJS middle +": 2500,
    "NodeJS senior -": 3050,
    "NodeJS senior": 3450,
    "NodeJS senior +": 3700,
    "NodeJS teamlead": 4300,
    "Front-end html/css middle -": 1100,
    "Front-end html/css middle": 1650,
    "Front-end html/css middle +": 1950,
    "Front-end middle": 2880,
    "Front-end senior": 3840,
    "NodeJS middle": 3520,
    "NodeJS senior": 4480,
    "Front-end html/css middle": 2240,
    "Mobile Dev Junior-": 850,
    "Mobile Dev Junior": 1400,
    "Mobile Dev Junior+": 1800,
    "Mobile Dev Middle-": 2050,
    "Mobile Dev Middle": 2600,
    "Mobile Dev Middle+": 3150,
    "Mobile Dev Senior-": 3450,
    "Mobile Dev Senior": 3900,
    "Mobile Dev Senior+": 4400,
    "Mobile Dev Teamlead": 5150,
    "Mobile dev middle": 3680,
    "Mobile dev senior": 5120,
    "Python junior": 950,
    "Python middle-": 2200,
    "Python middle": 2750,
    "Python middle+": 3300,
    "Python senior-": 3600,
    "Python senior": 4000,
    "Python senior+": 4550,
    "Python teamlead": 5000,
    "Golang junior": 950,
    "Golang middle-": 2200,
    "Golang middle": 2750,
    "Golang middle+": 3300,
    "Golang senior-": 3600,
    "Golang senior": 4150,
    "Golang senior+": 4550,
    "Golang teamlead": 5000,
    "Python middle": 3200,
    "Python senior": 4640,
    "Golang middle": 3200,
    "Golang senior": 4640,
    "Devops junior": 1100,
    "Devops middle-": 1950,
    "Devops middle": 2600,
    "Devops middle+": 3300,
    "Devops senior-": 3700,
    "Devops senior": 4150,
    "Devops senior+": 4450,
    "Devops teamlead": 5000,
    "Devops middle": 3200,
    "Devops senior": 4800,
    "Java middle": 3200,
    "Java senior": 4640,
    ".NET": 3200,
    ".NET Senior": 4640,
    "Security Specialist": 5600,
    "Lead Programmer Researcher / AI": 3437,
    "Programmer researcher / AI": 1980,
    "Senior data analyst / AI": 4097,
    "Middle data analyst / AI": 2564,
    "Junior data analyst / AI": 1753,
    "PM / AI": 2071,
    "Teamlead / AI": 2564,
    "Product analyst / AI": 1966,
    "Solution Architect consultant / AI": 2807,
    "Pr-manager junior": 950,
    "Pr-manager middle": 1900,
    "Pr-manager senior": 2350,
    "DevRel junior": 1100,
    "DevRel middle": 2050,
    "DevRel senior": 3150,
    "Copywriter middle": 1350,
    "Copywriter senior": 1900,
    "Photographer/Videographer": 3520,
    "Copywriter middle": 2160,
    "Copywriter senior": 3040,
    "Content Manager": 1600
}

#############################
# 2. Чтение Excel (с формулами как значениями)
#############################
def read_excel_data_only(file, sheet_name=0):
    file_data = file.read()
    wb = openpyxl.load_workbook(BytesIO(file_data), data_only=True)
    if isinstance(sheet_name, int):
        sheet = wb.worksheets[sheet_name]
    else:
        sheet = wb[sheet_name]
    data = list(sheet.values)
    df = pd.DataFrame(data)
    return df

#############################
# 3. Проверка грейдов
#############################
def check_grades(df, allowed_grades):
    unknown_grades = set()
    for col in df.columns:
        if len(df) > 2:
            third_row_value = df.iloc[2, col]
            if pd.notna(third_row_value):
                text = str(third_row_value).strip().lower()
                if text in ["inside", "outside"]:
                    grade_val = df.iloc[1, col]
                    if pd.notna(grade_val):
                        grade_str = str(grade_val).strip()
                        if grade_str not in allowed_grades:
                            unknown_grades.add(grade_str)
    return unknown_grades

#############################
# 4. Поиск столбца "Total cost"
#############################
def find_total_cost_column_name(df):
    for col in df.columns:
        cell_value = df.iloc[1, col]
        if pd.notna(cell_value) and str(cell_value).strip() == "Total cost":
            return col
    return None

#############################
# 5. Функция для обработки имени эпика
#############################
def process_function_name(epic_name):
    return "_".join(epic_name.split())

#############################
# 6. Жёстко берём столбцы F..Y как "оценочные"
#############################
def get_time_estimate_columns(df):
    grade_cols = []
    for col in range(5, 25):  # столбцы F..Y (индексы 5..24)
        if col < len(df.columns):
            third_row_value = df.iloc[2, col]
            if pd.notna(third_row_value):
                text = str(third_row_value).strip().lower()
                if text in ["inside", "outside"]:
                    grade_val = df.iloc[1, col]
                    if pd.notna(grade_val):
                        grade_str = str(grade_val).strip()
                        grade_cols.append((grade_str, col))
    return grade_cols

#############################
# 7. Функция для суммирования оценок (без отладочной печати)
#############################
def sum_estimates(row):
    total = 0
    for x in row:
        if x is None:
            continue
        elif isinstance(x, float) and math.isnan(x):
            continue
        elif isinstance(x, str) and x.strip().lower() == "null":
            continue
        else:
            try:
                total += float(x)
            except (TypeError, ValueError):
                continue
    return total if total != 0 else None

#############################
# 8. Обработка "с эпиками"
#############################
def process_with_epics(df):
    total_cost_col = find_total_cost_column_name(df)
    grade_cols = get_time_estimate_columns(df)
    start_row = 7  # Подберите под структуру

    df_subset = df.iloc[start_row:, [1, 2]].dropna(how='all').reset_index(drop=True)
    df_subset.columns = ['Feature', 'Details']

    summary_list = []
    custom_link_id_list = []
    parent_link_id_list = []
    issue_type_list = []
    total_cost_list = []
    function_name_list = []

    # Собираем уникальные грейды (по названию)
    unique_grades = set(g[0] for g in grade_cols)
    grade_values = {gname: [] for gname in unique_grades}

    current_custom_link_id = None
    current_function_name = None

    for idx, row in df_subset.iterrows():
        feature = row['Feature']
        detail = row['Details']
        original_row_index = idx + start_row

        cost_value = None
        if total_cost_col is not None and original_row_index < len(df):
            cost_value = df.iloc[original_row_index, total_cost_col]

        row_is_epic = (pd.notna(feature) and str(feature).strip() != "")
        row_is_ft   = (pd.notna(detail) and str(detail).strip() != "")

        if row_is_epic:
            custom_id = str(random.randint(100000, 999999))
            processed_fn = process_function_name(str(feature))
            summary_list.append(feature)
            issue_type_list.append("Epic")
            custom_link_id_list.append(custom_id)
            parent_link_id_list.append(None)
            total_cost_list.append(None)
            function_name_list.append(processed_fn)
            current_custom_link_id = custom_id
            current_function_name = processed_fn
            for gname in unique_grades:
                grade_values[gname].append(0.0)

        if row_is_ft:
            summary_list.append(detail)
            issue_type_list.append("ФТ")
            custom_link_id_list.append(None)
            parent_link_id_list.append(current_custom_link_id)
            total_cost_list.append(cost_value if pd.notna(cost_value) else None)
            function_name_list.append(current_function_name if current_function_name else None)
            for gname in unique_grades:
                grade_values[gname].append(0.0)
            row_in_csv = len(summary_list) - 1
            for (grade_name, col_index) in grade_cols:
                if original_row_index < len(df):
                    val = df.iloc[original_row_index, col_index]
                    if pd.notna(val) and float(val) != 0.0:
                        grade_values[grade_name][row_in_csv] += float(val)

    result_df = pd.DataFrame({
        'Summary': summary_list,
        'Custom Link ID': custom_link_id_list,
        'Parent Link ID': parent_link_id_list,
        'Issue Type': issue_type_list,
        'Total cost': total_cost_list,
        'Function name': function_name_list
    })

    for gname in unique_grades:
        values = [None if x == 0.0 else x for x in grade_values[gname]]
        result_df[gname] = values

    # Вычисляем столбец "Сумма времязатрат"
    grade_columns = list(unique_grades)
    result_df["Сумма времязатрат"] = result_df[grade_columns].apply(sum_estimates, axis=1)
    result_df.loc[result_df["Issue Type"] == "Epic", "Сумма времязатрат"] = None

    # Новая часть: вычисляем денежные затраты для каждого грейда
    for gname in unique_grades:
        cost_col_name = "[Cost] " + gname
        result_df[cost_col_name] = result_df[gname].apply(lambda x: x * hourly_rates[gname] if x is not None else None)

    # Отладка: выводим информацию по столбцам [Cost]
    cost_columns = ["[Cost] " + g for g in unique_grades]
    st.write("Отладка: столбцы [Cost]:")
    st.write(result_df[cost_columns])

    return result_df

#############################
# 9. Обработка "без эпиков"
#############################
def process_without_epics(df):
    total_cost_col = find_total_cost_column_name(df)
    grade_cols = get_time_estimate_columns(df)
    start_row = 7

    df_subset = df.iloc[start_row:, [1, 2]].dropna(how='all').reset_index(drop=True)
    df_subset.columns = ['Feature', 'Details']

    summary_list = []
    issue_type_list = []
    total_cost_list = []
    function_name_list = []

    unique_grades = set(g[0] for g in grade_cols)
    grade_values = {gname: [] for gname in unique_grades}

    current_function_name = None

    for idx, row in df_subset.iterrows():
        feature = row['Feature']
        detail = row['Details']
        original_row_index = idx + start_row

        cost_value = None
        if total_cost_col is not None and original_row_index < len(df):
            cost_value = df.iloc[original_row_index, total_cost_col]

        if pd.notna(feature) and str(feature).strip() != "":
            current_function_name = process_function_name(str(feature))

        if pd.notna(detail) and str(detail).strip() != "":
            summary_list.append(detail)
            issue_type_list.append("ФТ")
            total_cost_list.append(cost_value if pd.notna(cost_value) else None)
            function_name_list.append(current_function_name if current_function_name else None)
            for gname in unique_grades:
                grade_values[gname].append(0.0)
            row_in_csv = len(summary_list) - 1
            for (grade_name, col_index) in grade_cols:
                if original_row_index < len(df):
                    val = df.iloc[original_row_index, col_index]
                    if pd.notna(val) and float(val) != 0.0:
                        grade_values[grade_name][row_in_csv] += float(val)

    result_df = pd.DataFrame({
        'Summary': summary_list,
        'Issue Type': issue_type_list,
        'Total cost': total_cost_list,
        'Function name': function_name_list
    })

    for gname in unique_grades:
        values = [None if x == 0.0 else x for x in grade_values[gname]]
        result_df[gname] = values

    grade_columns = list(unique_grades)
    result_df["Сумма времязатрат"] = result_df[grade_columns].apply(sum_estimates, axis=1)
    result_df.loc[result_df["Issue Type"] == "Epic", "Сумма времязатрат"] = None

    for gname in unique_grades:
        cost_col_name = "[Cost] " + gname
        result_df[cost_col_name] = result_df[gname].apply(lambda x: x * hourly_rates[gname] if x is not None else None)

    cost_columns = ["[Cost] " + g for g in unique_grades]
    st.write("Отладка: столбцы [Cost] (без эпиков):")
    st.write(result_df[cost_columns])

    return result_df

#############################
# 10. Основной поток (Streamlit)
#############################
def main():
    st.title("Jira CSV Generator")

    processing_option = st.radio(
        "Выберите вариант обработки данных:",
        ("Импортировать Функции как Epic's", "Не импортировать Эпики")
    )

    uploaded_file = st.file_uploader("Загрузите Excel файл", type=["xlsx"])
    if uploaded_file:
        st.success("Файл успешно загружен!")
        df = read_excel_data_only(uploaded_file, sheet_name=0)

        unknown_grades = check_grades(df, list(hourly_rates.keys()))
        if unknown_grades:
            st.warning("Внимание! В смете присутствуют неизвестные грейды: " + ", ".join(unknown_grades))

        if processing_option == "Импортировать Функции как Epic's":
            result_df = process_with_epics(df)
        else:
            result_df = process_without_epics(df)

        st.dataframe(result_df)

        csv = result_df.to_csv(index=False).encode('utf-8')
        st.download_button(
            label="Скачать CSV файл",
            data=csv,
            file_name='Jira-Import.csv',
            mime='text/csv'
        )

    config_file_path = "Конфиг v2.txt"
    try:
        with open(config_file_path, 'r') as config_file:
            config_data = config_file.read()
        st.download_button(
            label="Скачать конфиг-файл для быстрого импорта",
            data=config_data,
            file_name='Jira-Import-Config.txt',
            mime='text/plain'
        )
    except FileNotFoundError:
        st.error(f"Файл {config_file_path} не найден. Убедитесь, что он загружен в репозиторий.")

if __name__ == "__main__":
    main()









# ВЕРСИЯ 0.4

# import streamlit as st
# import pandas as pd
# import random
# from io import BytesIO
# import openpyxl
# import math

# #############################
# # 1. Список допустимых грейдов
# #############################
# allowed_grades = [
#     "Продуктовая аналитика стажер",
#     "UX-аналитик junior",
#     "UX-аналитик middle -",
#     "UX-аналитик middle",
#     "UX-аналитик senior",
#     "Web/mobile-аналитик junior",
#     "Web/mobile-аналитик middle -",
#     "Web/mobile-аналитик middle",
#     "Web/mobile-аналитик middle +",
#     "Web/mobile-аналитик senior",
#     "Web/mobile-аналитик TL",
#     "BI/data-аналитик junior",
#     "BI/data-аналитик middle",
#     "BI/data-аналитик middle +",
#     "BI/data-аналитик senior",
#     "Product manager junior",
#     "Product manager middle",
#     "Product manager middle +",
#     "Product manager Senior",
#     "Web/mobile-analyst middle",
#     "Web/mobile-аналитик senior",
#     "Дизайнер стажер",
#     "Дизайнер junior-",
#     "Дизайнер junior",
#     "Дизайнер junior+",
#     "Дизайнер middle-",
#     "Дизайнер middle",
#     "Дизайнер middle+",
#     "Дизайнер senior",
#     "Дизайнер senior+",
#     "Art Director",
#     "Web middle",
#     "Web senior",
#     "Системный аналитик Стажер",
#     "Системный аналитик junior-",
#     "Системный аналитик junior",
#     "Системный аналитик junior+",
#     "Системный аналитик middle-",
#     "Системный аналитик middle",
#     "Системный аналитик middle+",
#     "Системный аналитик senior",
#     "Системный аналитик senior+",
#     "Системный аналитик lead",
#     "Проектировщик стажер",
#     "Проектировщик junior",
#     "Проектировщик middle",
#     "Проектировщик middle+",
#     "Проектировщик senior",
#     "Проектировщик senior+",
#     "Проектировщик lead",
#     "System analyst middle",
#     "System analyst senior",
#     "SEO",
#     "tech.writer",
#     "UX writer",
#     "PM стажер",
#     "PM intern",
#     "PM junior",
#     "PM junior+",
#     "PM middle",
#     "PM middle+",
#     "PM senior",
#     "PM senior+",
#     "GH",
#     "PMO",
#     "Bitrix junior",
#     "Bitrix junior+",
#     "Bitrix middle-",
#     "Bitrix middle",
#     "Bitrix middle+",
#     "Bitrix senior-",
#     "Bitrix senior",
#     "Bitrix senior+",
#     "Bitrix teamlead -",
#     "Bitrix teamlead",
#     "Bitrix teamlead +",
#     "Bitrix teamlead grouphead -",
#     "Bitrix teamlead grouphead",
#     "Bitrix teamlead grouphead +",
#     "Bitrix middle",
#     "Bitrix senior",
#     "Разработка стажер",
#     "Framework junior",
#     "Framework junior+",
#     "Framework middle-",
#     "Framework middle",
#     "Framework middle+",
#     "Framework senior-",
#     "Framework senior",
#     "Framework senior+",
#     "Framework teamlead -",
#     "Framework teamlead",
#     "Framework teamlead +",
#     "Framework teamlead grouphead -",
#     "Framework teamlead grouphead",
#     "Framework teamlead grouphead +",
#     "Framework middle",
#     "Framework senior",
#     "QA junior-",
#     "QA junior",
#     "QA junior +",
#     "QA middle-",
#     "QA middle",
#     "QA middle+",
#     "QA senior -",
#     "QA senior",
#     "QA Teamlead",
#     "AQA junior-",
#     "AQA junior",
#     "AQA junior +",
#     "AQA middle-",
#     "AQA middle",
#     "AQA middle+",
#     "AQA senior -",
#     "AQA senior",
#     "AQA teamlead",
#     "QA middle",
#     "QA senior",
#     "QA AT Middle",
#     "QA AT Senior",
#     "Front-end junior",
#     "Front-end junior +",
#     "Front-end middle -",
#     "Front-end middle",
#     "Front-end middle +",
#     "Front-end senior -",
#     "Front-end senior",
#     "Front-end senior +",
#     "Front-end teamlead",
#     "NodeJS junior",
#     "NodeJS junior +",
#     "NodeJS middle -",
#     "NodeJS middle",
#     "NodeJS middle +",
#     "NodeJS senior -",
#     "NodeJS senior",
#     "NodeJS senior +",
#     "NodeJS teamlead",
#     "Front-end html/css middle -",
#     "Front-end html/css middle",
#     "Front-end html/css middle +",
#     "Front-end middle",
#     "Front-end senior",
#     "NodeJS middle",
#     "NodeJS senior",
#     "Front-end html/css middle",
#     "Mobile Dev Junior-",
#     "Mobile Dev Junior",
#     "Mobile Dev Junior+",
#     "Mobile Dev Middle-",
#     "Mobile Dev Middle",
#     "Mobile Dev Middle+",
#     "Mobile Dev Senior-",
#     "Mobile Dev Senior",
#     "Mobile Dev Senior+",
#     "Mobile Dev Teamlead",
#     "Mobile dev middle",
#     "Mobile dev senior",
#     "Python junior",
#     "Python middle-",
#     "Python middle",
#     "Python middle+",
#     "Python senior-",
#     "Python senior",
#     "Python senior+",
#     "Python teamlead",
#     "Golang junior",
#     "Golang middle-",
#     "Golang middle",
#     "Golang middle+",
#     "Golang senior-",
#     "Golang senior",
#     "Golang senior+",
#     "Golang teamlead",
#     "Python middle",
#     "Python senior",
#     "Golang middle",
#     "Golang senior",
#     "Devops junior",
#     "Devops middle-",
#     "Devops middle",
#     "Devops middle+",
#     "Devops senior-",
#     "Devops senior",
#     "Devops senior+",
#     "Devops teamlead",
#     "Devops middle",
#     "Devops senior",
#     "Java middle",
#     "Java senior",
#     ".NET",
#     ".NET Senior",
#     "Security Specialist",
#     "Lead Programmer Researcher / AI",
#     "Programmer researcher / AI",
#     "Senior data analyst / AI",
#     "Middle data analyst / AI",
#     "Junior data analyst / AI",
#     "PM / AI",
#     "Teamlead / AI",
#     "Product analyst / AI",
#     "Solution Architect consultant / AI",
#     "Pr-manager junior",
#     "Pr-manager middle",
#     "Pr-manager senior",
#     "DevRel junior",
#     "DevRel middle",
#     "DevRel senior",
#     "Copywriter middle",
#     "Copywriter senior",
#     "Photographer/Videographer",
#     "Copywriter middle",
#     "Copywriter senior",
#     "Content Manager",
# ]

# #############################
# # 2. Чтение Excel (с формулами как значениями)
# #############################
# def read_excel_data_only(file, sheet_name=0):
#     file_data = file.read()
#     wb = openpyxl.load_workbook(BytesIO(file_data), data_only=True)
#     if isinstance(sheet_name, int):
#         sheet = wb.worksheets[sheet_name]
#     else:
#         sheet = wb[sheet_name]
#     data = list(sheet.values)
#     df = pd.DataFrame(data)
#     return df

# #############################
# # 3. Проверка грейдов
# #############################
# def check_grades(df, allowed_grades):
#     unknown_grades = set()
#     for col in df.columns:
#         if len(df) > 2:
#             third_row_value = df.iloc[2, col]
#             if pd.notna(third_row_value):
#                 text = str(third_row_value).strip().lower()
#                 if text in ["inside", "outside"]:
#                     grade_val = df.iloc[1, col]
#                     if pd.notna(grade_val):
#                         grade_str = str(grade_val).strip()
#                         if grade_str not in allowed_grades:
#                             unknown_grades.add(grade_str)
#     return unknown_grades

# #############################
# # 4. Поиск столбца "Total cost"
# #############################
# def find_total_cost_column_name(df):
#     for col in df.columns:
#         cell_value = df.iloc[1, col]
#         if pd.notna(cell_value) and str(cell_value).strip() == "Total cost":
#             return col
#     return None

# #############################
# # 5. Функция для обработки имени эпика
# #############################
# def process_function_name(epic_name):
#     return "_".join(epic_name.split())

# #############################
# # 6. Жёстко берём столбцы F..Y как "оценочные"
# #############################
# def get_time_estimate_columns(df):
#     grade_cols = []
#     for col in range(5, 25):  # столбцы F..Y (индексы 5..24)
#         if col < len(df.columns):
#             third_row_value = df.iloc[2, col]
#             if pd.notna(third_row_value):
#                 text = str(third_row_value).strip().lower()
#                 if text in ["inside", "outside"]:
#                     grade_val = df.iloc[1, col]
#                     if pd.notna(grade_val):
#                         grade_str = str(grade_val).strip()
#                         grade_cols.append((grade_str, col))
#     return grade_cols

# #############################
# # 7. Функция для суммирования оценок с учётом NaN/NULL
# #############################
# def sum_estimates_debug(row):
#     st.write("Отладка, строка:", row.tolist())
#     total = 0
#     for x in row:
#         if x is None:
#             # None трактуем как 0
#             continue
#         elif isinstance(x, float) and math.isnan(x):
#             # NaN трактуем как 0
#             continue
#         elif isinstance(x, str) and x.strip().lower() == "null":
#             # Строка "NULL" тоже трактуем как 0
#             continue
#         else:
#             # Пытаемся привести к float
#             try:
#                 total += float(x)
#             except (TypeError, ValueError):
#                 # Если что-то пошло не так, прибавляем 0
#                 continue
#     st.write("Отладка, сумма:", total)
#     return total if total != 0 else None

# #############################
# # 8. Обработка "с эпиками"
# #############################
# def process_with_epics(df):
#     total_cost_col = find_total_cost_column_name(df)
#     grade_cols = get_time_estimate_columns(df)
#     start_row = 7  # Подберите под структуру

#     df_subset = df.iloc[start_row:, [1, 2]].dropna(how='all').reset_index(drop=True)
#     df_subset.columns = ['Feature', 'Details']

#     summary_list = []
#     custom_link_id_list = []
#     parent_link_id_list = []
#     issue_type_list = []
#     total_cost_list = []
#     function_name_list = []

#     # Собираем уникальные грейды (по названию)
#     unique_grades = set(g[0] for g in grade_cols)
#     grade_values = {gname: [] for gname in unique_grades}

#     current_custom_link_id = None
#     current_function_name = None

#     for idx, row in df_subset.iterrows():
#         feature = row['Feature']
#         detail = row['Details']
#         original_row_index = idx + start_row

#         cost_value = None
#         if total_cost_col is not None and original_row_index < len(df):
#             cost_value = df.iloc[original_row_index, total_cost_col]

#         row_is_epic = (pd.notna(feature) and str(feature).strip() != "")
#         row_is_ft   = (pd.notna(detail) and str(detail).strip() != "")

#         if row_is_epic:
#             custom_id = str(random.randint(100000, 999999))
#             processed_fn = process_function_name(str(feature))
#             summary_list.append(feature)
#             issue_type_list.append("Epic")
#             custom_link_id_list.append(custom_id)
#             parent_link_id_list.append(None)
#             total_cost_list.append(None)
#             function_name_list.append(processed_fn)
#             current_custom_link_id = custom_id
#             current_function_name = processed_fn
#             # Для каждой строки эпика добавляем для всех уникальных грейдов начальное значение 0.0
#             for gname in unique_grades:
#                 grade_values[gname].append(0.0)

#         if row_is_ft:
#             summary_list.append(detail)
#             issue_type_list.append("ФТ")
#             custom_link_id_list.append(None)
#             parent_link_id_list.append(current_custom_link_id)
#             total_cost_list.append(cost_value if pd.notna(cost_value) else None)
#             function_name_list.append(current_function_name if current_function_name else None)
#             for gname in unique_grades:
#                 grade_values[gname].append(0.0)
#             row_in_csv = len(summary_list) - 1
#             for (grade_name, col_index) in grade_cols:
#                 if original_row_index < len(df):
#                     val = df.iloc[original_row_index, col_index]
#                     if pd.notna(val) and float(val) != 0.0:
#                         grade_values[grade_name][row_in_csv] += float(val)

#     result_df = pd.DataFrame({
#         'Summary': summary_list,
#         'Custom Link ID': custom_link_id_list,
#         'Parent Link ID': parent_link_id_list,
#         'Issue Type': issue_type_list,
#         'Total cost': total_cost_list,
#         'Function name': function_name_list
#     })

#     # Преобразуем 0.0 в None для столбцов с оценками
#     for gname in unique_grades:
#         values = [None if x == 0.0 else x for x in grade_values[gname]]
#         result_df[gname] = values

#     grade_columns = list(unique_grades)
#     st.write("Отладка: данные по грейдам (с эпиками):")
#     st.write(result_df[grade_columns])

#     # Используем отладочную функцию суммирования
#     result_df["Сумма времязатрат"] = result_df[grade_columns].apply(sum_estimates_debug, axis=1)
#     result_df.loc[result_df["Issue Type"] == "Epic", "Сумма времязатрат"] = None

#     st.write("### Отладочная информация (с эпиками)")
#     st.write(f"Итоговое количество строк: {len(result_df)}")
#     st.write(f"Количество строк в summary_list: {len(summary_list)}")
#     for gname in unique_grades:
#         st.write(f"Грейд '{gname}': {len(grade_values[gname])} записей")

#     return result_df

# #############################
# # 9. Обработка "без эпиков"
# #############################
# def process_without_epics(df):
#     total_cost_col = find_total_cost_column_name(df)
#     grade_cols = get_time_estimate_columns(df)
#     start_row = 7

#     df_subset = df.iloc[start_row:, [1, 2]].dropna(how='all').reset_index(drop=True)
#     df_subset.columns = ['Feature', 'Details']

#     summary_list = []
#     issue_type_list = []
#     total_cost_list = []
#     function_name_list = []

#     unique_grades = set(g[0] for g in grade_cols)
#     grade_values = {gname: [] for gname in unique_grades}

#     current_function_name = None

#     for idx, row in df_subset.iterrows():
#         feature = row['Feature']
#         detail = row['Details']
#         original_row_index = idx + start_row

#         cost_value = None
#         if total_cost_col is not None and original_row_index < len(df):
#             cost_value = df.iloc[original_row_index, total_cost_col]

#         if pd.notna(feature) and str(feature).strip() != "":
#             current_function_name = process_function_name(str(feature))

#         if pd.notna(detail) and str(detail).strip() != "":
#             summary_list.append(detail)
#             issue_type_list.append("ФТ")
#             total_cost_list.append(cost_value if pd.notna(cost_value) else None)
#             function_name_list.append(current_function_name if current_function_name else None)
#             for gname in unique_grades:
#                 grade_values[gname].append(0.0)
#             row_in_csv = len(summary_list) - 1
#             for (grade_name, col_index) in grade_cols:
#                 if original_row_index < len(df):
#                     val = df.iloc[original_row_index, col_index]
#                     if pd.notna(val) and float(val) != 0.0:
#                         grade_values[grade_name][row_in_csv] += float(val)

#     result_df = pd.DataFrame({
#         'Summary': summary_list,
#         'Issue Type': issue_type_list,
#         'Total cost': total_cost_list,
#         'Function name': function_name_list
#     })

#     for gname in unique_grades:
#         values = [None if x == 0.0 else x for x in grade_values[gname]]
#         result_df[gname] = values

#     grade_columns = list(unique_grades)
#     st.write("Отладка: данные по грейдам (без эпиков):")
#     st.write(result_df[grade_columns])
#     result_df["Сумма времязатрат"] = result_df[grade_columns].apply(sum_estimates_debug, axis=1)
#     result_df.loc[result_df["Issue Type"] == "Epic", "Сумма времязатрат"] = None

#     st.write("### Отладочная информация (без эпиков)")
#     st.write(f"Итоговое количество строк: {len(result_df)}")
#     st.write(f"Количество строк в summary_list: {len(summary_list)}")
#     for gname in unique_grades:
#         st.write(f"Грейд '{gname}': {len(grade_values[gname])} записей")

#     return result_df

# #############################
# # 10. Основной поток (Streamlit)
# #############################
# def main():
#     st.title("Jira CSV Generator")

#     processing_option = st.radio(
#         "Выберите вариант обработки данных:",
#         ("Импортировать Функции как Epic's", "Не импортировать Эпики")
#     )

#     uploaded_file = st.file_uploader("Загрузите Excel файл", type=["xlsx"])
#     if uploaded_file:
#         st.success("Файл успешно загружен!")
#         df = read_excel_data_only(uploaded_file, sheet_name=0)

#         unknown_grades = check_grades(df, allowed_grades)
#         if unknown_grades:
#             st.warning("Внимание! В смете присутствуют неизвестные грейды: " + ", ".join(unknown_grades))

#         if processing_option == "Импортировать Функции как Epic's":
#             result_df = process_with_epics(df)
#         else:
#             result_df = process_without_epics(df)

#         st.dataframe(result_df)

#         csv = result_df.to_csv(index=False).encode('utf-8')
#         st.download_button(
#             label="Скачать CSV файл",
#             data=csv,
#             file_name='Jira-Import.csv',
#             mime='text/csv'
#         )

#     config_file_path = "Конфиг v2.txt"
#     try:
#         with open(config_file_path, 'r') as config_file:
#             config_data = config_file.read()
#         st.download_button(
#             label="Скачать конфиг-файл для быстрого импорта",
#             data=config_data,
#             file_name='Jira-Import-Config.txt',
#             mime='text/plain'
#         )
#     except FileNotFoundError:
#         st.error(f"Файл {config_file_path} не найден. Убедитесь, что он загружен в репозиторий.")


# if __name__ == "__main__":
#     main()











# ПЕРВАЯ РАБОЧАЯ ВЕРСИЯ ОТ 17 МАРТА. РАБОТАЕТ СЛОЖЕНИЕ ВРЕМЯЗАТРАТ В СТОЛБЦАХ С ГРЕЙДАМИ, ВСТРЕЧАЮЩИМИСЯ 2+ РАЗ

# import streamlit as st
# import pandas as pd
# import random
# from io import BytesIO
# import openpyxl
# from collections import defaultdict

# #############################
# # 1. Список допустимых грейдов
# #############################
# allowed_grades = [
#     "Продуктовая аналитика стажер",
#     "UX-аналитик junior",
#     "UX-аналитик middle -",
#     "UX-аналитик middle",
#     "UX-аналитик senior",
#     "Web/mobile-аналитик junior",
#     "Web/mobile-аналитик middle -",
#     "Web/mobile-аналитик middle",
#     "Web/mobile-аналитик middle +",
#     "Web/mobile-аналитик senior",
#     "Web/mobile-аналитик TL",
#     "BI/data-аналитик junior",
#     "BI/data-аналитик middle",
#     "BI/data-аналитик middle +",
#     "BI/data-аналитик senior",
#     "Product manager junior",
#     "Product manager middle",
#     "Product manager middle +",
#     "Product manager Senior",
#     "Web/mobile-analyst middle",
#     "Web/mobile-analyst senior",
#     "Дизайнер стажер",
#     "Дизайнер junior-",
#     "Дизайнер junior",
#     "Дизайнер junior+",
#     "Дизайнер middle-",
#     "Дизайнер middle",
#     "Дизайнер middle+",
#     "Дизайнер senior",
#     "Дизайнер senior+",
#     "Art Director",
#     "Web middle",
#     "Web senior",
#     "Системный аналитик Стажер",
#     "Системный аналитик junior-",
#     "Системный аналитик junior",
#     "Системный аналитик junior+",
#     "Системный аналитик middle-",
#     "Системный аналитик middle",
#     "Системный аналитик middle+",
#     "Системный аналитик senior",
#     "Системный аналитик senior+",
#     "Системный аналитик lead",
#     "Проектировщик стажер",
#     "Проектировщик junior",
#     "Проектировщик middle",
#     "Проектировщик middle+",
#     "Проектировщик senior",
#     "Проектировщик senior+",
#     "Проектировщик lead",
#     "System analyst middle",
#     "System analyst senior",
#     "SEO",
#     "tech.writer",
#     "UX writer",
#     "PM стажер",
#     "PM intern",
#     "PM junior",
#     "PM junior+",
#     "PM middle",
#     "PM middle+",
#     "PM senior",
#     "PM senior+",
#     "GH",
#     "PMO",
#     "Bitrix junior",
#     "Bitrix junior+",
#     "Bitrix middle-",
#     "Bitrix middle",
#     "Bitrix middle+",
#     "Bitrix senior-",
#     "Bitrix senior",
#     "Bitrix senior+",
#     "Bitrix teamlead -",
#     "Bitrix teamlead",
#     "Bitrix teamlead +",
#     "Bitrix teamlead grouphead -",
#     "Bitrix teamlead grouphead",
#     "Bitrix teamlead grouphead +",
#     "Bitrix middle",
#     "Bitrix senior",
#     "Разработка стажер",
#     "Framework junior",
#     "Framework junior+",
#     "Framework middle-",
#     "Framework middle",
#     "Framework middle+",
#     "Framework senior-",
#     "Framework senior",
#     "Framework senior+",
#     "Framework teamlead -",
#     "Framework teamlead",
#     "Framework teamlead +",
#     "Framework teamlead grouphead -",
#     "Framework teamlead grouphead",
#     "Framework teamlead grouphead +",
#     "Framework middle",
#     "Framework senior",
#     "QA junior-",
#     "QA junior",
#     "QA junior +",
#     "QA middle-",
#     "QA middle",
#     "QA middle+",
#     "QA senior -",
#     "QA senior",
#     "QA Teamlead",
#     "AQA junior-",
#     "AQA junior",
#     "AQA junior +",
#     "AQA middle-",
#     "AQA middle",
#     "AQA middle+",
#     "AQA senior -",
#     "AQA senior",
#     "AQA teamlead",
#     "QA middle",
#     "QA senior",
#     "QA AT Middle",
#     "QA AT Senior",
#     "Front-end junior",
#     "Front-end junior +",
#     "Front-end middle -",
#     "Front-end middle",
#     "Front-end middle +",
#     "Front-end senior -",
#     "Front-end senior",
#     "Front-end senior +",
#     "Front-end teamlead",
#     "NodeJS junior",
#     "NodeJS junior +",
#     "NodeJS middle -",
#     "NodeJS middle",
#     "NodeJS middle +",
#     "NodeJS senior -",
#     "NodeJS senior",
#     "NodeJS senior +",
#     "NodeJS teamlead",
#     "Front-end html/css middle -",
#     "Front-end html/css middle",
#     "Front-end html/css middle +",
#     "Front-end middle",
#     "Front-end senior",
#     "NodeJS middle",
#     "NodeJS senior",
#     "Front-end html/css middle",
#     "Mobile Dev Junior-",
#     "Mobile Dev Junior",
#     "Mobile Dev Junior+",
#     "Mobile Dev Middle-",
#     "Mobile Dev Middle",
#     "Mobile Dev Middle+",
#     "Mobile Dev Senior-",
#     "Mobile Dev Senior",
#     "Mobile Dev Senior+",
#     "Mobile Dev Teamlead",
#     "Mobile dev middle",
#     "Mobile dev senior",
#     "Python junior",
#     "Python middle-",
#     "Python middle",
#     "Python middle+",
#     "Python senior-",
#     "Python senior",
#     "Python senior+",
#     "Python teamlead",
#     "Golang junior",
#     "Golang middle-",
#     "Golang middle",
#     "Golang middle+",
#     "Golang senior-",
#     "Golang senior",
#     "Golang senior+",
#     "Golang teamlead",
#     "Python middle",
#     "Python senior",
#     "Golang middle",
#     "Golang senior",
#     "Devops junior",
#     "Devops middle-",
#     "Devops middle",
#     "Devops middle+",
#     "Devops senior-",
#     "Devops senior",
#     "Devops senior+",
#     "Devops teamlead",
#     "Devops middle",
#     "Devops senior",
#     "Java middle",
#     "Java senior",
#     ".NET",
#     ".NET Senior",
#     "Security Specialist",
#     "Lead Programmer Researcher / AI",
#     "Programmer researcher / AI",
#     "Senior data analyst / AI",
#     "Middle data analyst / AI",
#     "Junior data analyst / AI",
#     "PM / AI",
#     "Teamlead / AI",
#     "Product analyst / AI",
#     "Solution Architect consultant / AI",
#     "Pr-manager junior",
#     "Pr-manager middle",
#     "Pr-manager senior",
#     "DevRel junior",
#     "DevRel middle",
#     "DevRel senior",
#     "Copywriter middle",
#     "Copywriter senior",
#     "Photographer/Videographer",
#     "Copywriter middle",
#     "Copywriter senior",
#     "Content Manager",
# ]

# #############################
# # 2. Чтение Excel (с формулами как значениями)
# #############################
# def read_excel_data_only(file, sheet_name=0):
#     file_data = file.read()
#     wb = openpyxl.load_workbook(BytesIO(file_data), data_only=True)
#     if isinstance(sheet_name, int):
#         sheet = wb.worksheets[sheet_name]
#     else:
#         sheet = wb[sheet_name]
#     data = list(sheet.values)
#     df = pd.DataFrame(data)
#     return df

# #############################
# # 3. Проверка грейдов
# #############################
# def check_grades(df, allowed_grades):
#     unknown_grades = set()
#     for col in df.columns:
#         if len(df) > 2:
#             third_row_value = df.iloc[2, col]
#             if pd.notna(third_row_value):
#                 text = str(third_row_value).strip().lower()
#                 if text in ["inside", "outside"]:
#                     grade_val = df.iloc[1, col]
#                     if pd.notna(grade_val):
#                         grade_str = str(grade_val).strip()
#                         if grade_str not in allowed_grades:
#                             unknown_grades.add(grade_str)
#     return unknown_grades

# #############################
# # 4. Поиск столбца "Total cost"
# #############################
# def find_total_cost_column_name(df):
#     for col in df.columns:
#         cell_value = df.iloc[1, col]
#         if pd.notna(cell_value) and str(cell_value).strip() == "Total cost":
#             return col
#     return None

# #############################
# # 5. Функция для обработки имени эпика
# #############################
# def process_function_name(epic_name):
#     # Заменяем пробелы на нижние подчёркивания
#     return "_".join(epic_name.split())

# #############################
# # 6. Жёстко берём столбцы F..Y как "оценочные"
# #############################
# def get_time_estimate_columns(df):
#     """
#     Возвращает список кортежей (grade_name, col_index) для столбцов F..Y (индексы 5..24),
#     где во 2-й строке (index=1) указано название грейда,
#     а в 3-й строке (index=2) написано 'Inside' или 'Outside'.

#     ВАЖНО: Если грейд 'Дизайнер middle' встречается в двух столбцах,
#     в grade_cols будет [( 'Дизайнер middle', 9 ), ( 'Дизайнер middle', 10 )] и т.д.
#     """
#     grade_cols = []
#     for col in range(5, 25):  # F..Y
#         if col < len(df.columns):
#             third_row_value = df.iloc[2, col]
#             if pd.notna(third_row_value):
#                 text = str(third_row_value).strip().lower()
#                 if text in ["inside", "outside"]:
#                     grade_val = df.iloc[1, col]
#                     if pd.notna(grade_val):
#                         grade_str = str(grade_val).strip()
#                         grade_cols.append((grade_str, col))
#     return grade_cols

# #############################
# # 7. Обработка "с эпиками"
# #############################
# def process_with_epics(df):
#     total_cost_col = find_total_cost_column_name(df)
#     grade_cols = get_time_estimate_columns(df)  # список (grade_name, col_index)
#     start_row = 7  # Подберите под структуру

#     # Подготовка df_subset
#     df_subset = df.iloc[start_row:, [1, 2]].dropna(how='all').reset_index(drop=True)
#     df_subset.columns = ['Feature', 'Details']

#     # Основные списки
#     summary_list = []
#     custom_link_id_list = []
#     parent_link_id_list = []
#     issue_type_list = []
#     total_cost_list = []
#     function_name_list = []

#     # Собираем УНИКАЛЬНЫЕ названия грейдов,
#     # чтобы потом завести под каждый грейд ОДИН столбец.
#     unique_grades = set(g[0] for g in grade_cols)

#     # Вместо "grade_values[grade_name] = []", используем defaultdict(list) или dict c инициализацией
#     grade_values = {}
#     for gname in unique_grades:
#         grade_values[gname] = []  # один список на каждый уникальный грейд

#     current_custom_link_id = None
#     current_function_name = None

#     # -- Основной цикл --
#     for idx, row in df_subset.iterrows():
#         feature = row['Feature']
#         detail = row['Details']
#         original_row_index = idx + start_row

#         # Смотрим на Total cost
#         cost_value = None
#         if total_cost_col is not None and original_row_index < len(df):
#             cost_value = df.iloc[original_row_index, total_cost_col]

#         row_is_epic = (pd.notna(feature) and str(feature).strip() != "")
#         row_is_ft   = (pd.notna(detail)  and str(detail).strip()  != "")

#         # Если строка-эпик
#         if row_is_epic:
#             custom_id = str(random.randint(100000, 999999))
#             processed_fn = process_function_name(str(feature))

#             summary_list.append(feature)
#             issue_type_list.append("Epic")
#             custom_link_id_list.append(custom_id)
#             parent_link_id_list.append(None)
#             total_cost_list.append(None)
#             function_name_list.append(processed_fn)

#             current_custom_link_id = custom_id
#             current_function_name = processed_fn

#             # Для каждого уникального грейда добавляем НОВУЮ строку со значением 0.0 (или None).
#             # Предположим, что для эпика = 0.0
#             for gname in unique_grades:
#                 grade_values[gname].append(0.0)

#         # Если строка-ФТ
#         if row_is_ft:
#             summary_list.append(detail)
#             issue_type_list.append("ФТ")
#             custom_link_id_list.append(None)
#             parent_link_id_list.append(current_custom_link_id)
#             total_cost_list.append(cost_value if pd.notna(cost_value) else None)
#             function_name_list.append(current_function_name if current_function_name else None)

#             # Для каждого уникального грейда добавляем НОВУЮ строку (0.0 по умолчанию)
#             for gname in unique_grades:
#                 grade_values[gname].append(0.0)

#             # Теперь, т.к. мы только что добавили строку для ФТ, она стала ПОСЛЕДНЕЙ в списке.
#             # Её индекс:
#             row_in_csv = len(summary_list) - 1

#             # Суммируем значения из всех столбцов, где grade_name совпадает
#             # Если у нас 2 столбца с "Дизайнер middle", мы добавим их значения в один и тот же элемент
#             for (grade_name, col_index) in grade_cols:
#                 if original_row_index < len(df):
#                     val = df.iloc[original_row_index, col_index]
#                     if pd.notna(val) and float(val) != 0.0:
#                         # Прибавляем к уже имеющемуся значению
#                         grade_values[grade_name][row_in_csv] += float(val)

#     # -- Формируем итоговый DataFrame --
#     result_df = pd.DataFrame({
#         'Summary': summary_list,
#         'Custom Link ID': custom_link_id_list,
#         'Parent Link ID': parent_link_id_list,
#         'Issue Type': issue_type_list,
#         'Total cost': total_cost_list,
#         'Function name': function_name_list
#     })

#     # Превращаем нули в None, если нужно
#     # (или можно оставить 0.0, как вам удобнее)
#     for gname in unique_grades:
#         # Если хотим скрыть нули:
#         # values = [None if x == 0.0 else x for x in grade_values[gname]]
#         # Если хотим оставить нули, используем без преобразования
#         values = grade_values[gname]
#         result_df[gname] = values

#     return result_df

# #############################
# # 8. Обработка "без эпиков"
# #############################
# def process_without_epics(df):
#     total_cost_col = find_total_cost_column_name(df)
#     grade_cols = get_time_estimate_columns(df)
#     start_row = 7

#     df_subset = df.iloc[start_row:, [1, 2]].dropna(how='all').reset_index(drop=True)
#     df_subset.columns = ['Feature', 'Details']

#     summary_list = []
#     issue_type_list = []
#     total_cost_list = []
#     function_name_list = []

#     # Собираем уникальные грейды
#     unique_grades = set(g[0] for g in grade_cols)
#     grade_values = {}
#     for gname in unique_grades:
#         grade_values[gname] = []

#     current_function_name = None

#     for idx, row in df_subset.iterrows():
#         feature = row['Feature']
#         detail = row['Details']
#         original_row_index = idx + start_row

#         cost_value = None
#         if total_cost_col is not None and original_row_index < len(df):
#             cost_value = df.iloc[original_row_index, total_cost_col]

#         row_is_epic = (pd.notna(feature) and str(feature).strip() != "")
#         row_is_ft   = (pd.notna(detail)  and str(detail).strip()  != "")

#         # Если встретили эпик — просто запоминаем название, но не создаём строку
#         if row_is_epic:
#             current_function_name = process_function_name(str(feature))

#         # Если это ФТ
#         if row_is_ft:
#             summary_list.append(detail)
#             issue_type_list.append("ФТ")
#             total_cost_list.append(cost_value if pd.notna(cost_value) else None)
#             function_name_list.append(current_function_name if current_function_name else None)

#             # Добавляем новую строку (значение 0.0) для каждого грейда
#             for gname in unique_grades:
#                 grade_values[gname].append(0.0)

#             row_in_csv = len(summary_list) - 1

#             # Суммируем значения
#             for (grade_name, col_index) in grade_cols:
#                 if original_row_index < len(df):
#                     val = df.iloc[original_row_index, col_index]
#                     if pd.notna(val) and float(val) != 0.0:
#                         grade_values[grade_name][row_in_csv] += float(val)

#     result_df = pd.DataFrame({
#         'Summary': summary_list,
#         'Issue Type': issue_type_list,
#         'Total cost': total_cost_list,
#         'Function name': function_name_list
#     })

#     # Превращаем нули в None (или оставляем как есть)
#     for gname in unique_grades:
#         values = grade_values[gname]
#         # values = [None if x == 0.0 else x for x in values]  # если нужно скрывать нули
#         result_df[gname] = values

#     return result_df

# #############################
# # 9. Основной поток (Streamlit)
# #############################
# def main():
#     st.title("Jira CSV Generator")

#     processing_option = st.radio(
#         "Выберите вариант обработки данных:",
#         ("Импортировать Функции как Epic's", "Не импортировать Эпики")
#     )

#     uploaded_file = st.file_uploader("Загрузите Excel файл", type=["xlsx"])
#     if uploaded_file:
#         st.success("Файл успешно загружен!")
#         df = read_excel_data_only(uploaded_file, sheet_name=0)

#         # Проверяем грейды
#         unknown_grades = check_grades(df, allowed_grades)
#         if unknown_grades:
#             st.warning(
#                 "Внимание! В смете присутствуют неизвестные грейды: " + ", ".join(unknown_grades)
#             )

#         if processing_option == "Импортировать Функции как Epic's":
#             result_df = process_with_epics(df)
#         else:
#             result_df = process_without_epics(df)

#         st.dataframe(result_df)

#         # Скачивание CSV
#         csv = result_df.to_csv(index=False).encode('utf-8')
#         st.download_button(
#             label="Скачать CSV файл",
#             data=csv,
#             file_name='Jira-Import.csv',
#             mime='text/csv'
#         )

#     # Кнопка для скачивания конфиг-файла
#     config_file_path = "Конфиг v2.txt"
#     try:
#         with open(config_file_path, 'r') as config_file:
#             config_data = config_file.read()
#         st.download_button(
#             label="Скачать конфиг-файл для быстрого импорта",
#             data=config_data,
#             file_name='Jira-Import-Config.txt',
#             mime='text/plain'
#         )
#     except FileNotFoundError:
#         st.error(f"Файл {config_file_path} не найден. Убедитесь, что он загружен в репозиторий.")


# if __name__ == "__main__":
#     main()









# ПРЕДЫДУЩАЯ ВЕРСИЯ ОТ 17 МАРТА

# import streamlit as st
# import pandas as pd
# import random
# from io import BytesIO
# import openpyxl

# #############################
# # 1. Список допустимых грейдов
# #############################
# allowed_grades = [
#     "Продуктовая аналитика стажер",
#     "UX-аналитик junior",
#     "UX-аналитик middle -",
#     "UX-аналитик middle",
#     "UX-аналитик senior",
#     "Web/mobile-аналитик junior",
#     "Web/mobile-аналитик middle -",
#     "Web/mobile-аналитик middle",
#     "Web/mobile-аналитик middle +",
#     "Web/mobile-аналитик senior",
#     "Web/mobile-аналитик TL",
#     "BI/data-аналитик junior",
#     "BI/data-аналитик middle",
#     "BI/data-аналитик middle +",
#     "BI/data-аналитик senior",
#     "Product manager junior",
#     "Product manager middle",
#     "Product manager middle +",
#     "Product manager Senior",
#     "Web/mobile-analyst middle",
#     "Web/mobile-analyst senior",
#     "Дизайнер стажер",
#     "Дизайнер junior-",
#     "Дизайнер junior",
#     "Дизайнер junior+",
#     "Дизайнер middle-",
#     "Дизайнер middle",
#     "Дизайнер middle+",
#     "Дизайнер senior",
#     "Дизайнер senior+",
#     "Art Director",
#     "Web middle",
#     "Web senior",
#     "Системный аналитик Стажер",
#     "Системный аналитик junior-",
#     "Системный аналитик junior",
#     "Системный аналитик junior+",
#     "Системный аналитик middle-",
#     "Системный аналитик middle",
#     "Системный аналитик middle+",
#     "Системный аналитик senior",
#     "Системный аналитик senior+",
#     "Системный аналитик lead",
#     "Проектировщик стажер",
#     "Проектировщик junior",
#     "Проектировщик middle",
#     "Проектировщик middle+",
#     "Проектировщик senior",
#     "Проектировщик senior+",
#     "Проектировщик lead",
#     "System analyst middle",
#     "System analyst senior",
#     "SEO",
#     "tech.writer",
#     "UX writer",
#     "PM стажер",
#     "PM intern",
#     "PM junior",
#     "PM junior+",
#     "PM middle",
#     "PM middle+",
#     "PM senior",
#     "PM senior+",
#     "GH",
#     "PMO",
#     "Bitrix junior",
#     "Bitrix junior+",
#     "Bitrix middle-",
#     "Bitrix middle",
#     "Bitrix middle+",
#     "Bitrix senior-",
#     "Bitrix senior",
#     "Bitrix senior+",
#     "Bitrix teamlead -",
#     "Bitrix teamlead",
#     "Bitrix teamlead +",
#     "Bitrix teamlead grouphead -",
#     "Bitrix teamlead grouphead",
#     "Bitrix teamlead grouphead +",
#     "Bitrix middle",
#     "Bitrix senior",
#     "Разработка стажер",
#     "Framework junior",
#     "Framework junior+",
#     "Framework middle-",
#     "Framework middle",
#     "Framework middle+",
#     "Framework senior-",
#     "Framework senior",
#     "Framework senior+",
#     "Framework teamlead -",
#     "Framework teamlead",
#     "Framework teamlead +",
#     "Framework teamlead grouphead -",
#     "Framework teamlead grouphead",
#     "Framework teamlead grouphead +",
#     "Framework middle",
#     "Framework senior",
#     "QA junior-",
#     "QA junior",
#     "QA junior +",
#     "QA middle-",
#     "QA middle",
#     "QA middle+",
#     "QA senior -",
#     "QA senior",
#     "QA Teamlead",
#     "AQA junior-",
#     "AQA junior",
#     "AQA junior +",
#     "AQA middle-",
#     "AQA middle",
#     "AQA middle+",
#     "AQA senior -",
#     "AQA senior",
#     "AQA teamlead",
#     "QA middle",
#     "QA senior",
#     "QA AT Middle",
#     "QA AT Senior",
#     "Front-end junior",
#     "Front-end junior +",
#     "Front-end middle -",
#     "Front-end middle",
#     "Front-end middle +",
#     "Front-end senior -",
#     "Front-end senior",
#     "Front-end senior +",
#     "Front-end teamlead",
#     "NodeJS junior",
#     "NodeJS junior +",
#     "NodeJS middle -",
#     "NodeJS middle",
#     "NodeJS middle +",
#     "NodeJS senior -",
#     "NodeJS senior",
#     "NodeJS senior +",
#     "NodeJS teamlead",
#     "Front-end html/css middle -",
#     "Front-end html/css middle",
#     "Front-end html/css middle +",
#     "Front-end middle",
#     "Front-end senior",
#     "NodeJS middle",
#     "NodeJS senior",
#     "Front-end html/css middle",
#     "Mobile Dev Junior-",
#     "Mobile Dev Junior",
#     "Mobile Dev Junior+",
#     "Mobile Dev Middle-",
#     "Mobile Dev Middle",
#     "Mobile Dev Middle+",
#     "Mobile Dev Senior-",
#     "Mobile Dev Senior",
#     "Mobile Dev Senior+",
#     "Mobile Dev Teamlead",
#     "Mobile dev middle",
#     "Mobile dev senior",
#     "Python junior",
#     "Python middle-",
#     "Python middle",
#     "Python middle+",
#     "Python senior-",
#     "Python senior",
#     "Python senior+",
#     "Python teamlead",
#     "Golang junior",
#     "Golang middle-",
#     "Golang middle",
#     "Golang middle+",
#     "Golang senior-",
#     "Golang senior",
#     "Golang senior+",
#     "Golang teamlead",
#     "Python middle",
#     "Python senior",
#     "Golang middle",
#     "Golang senior",
#     "Devops junior",
#     "Devops middle-",
#     "Devops middle",
#     "Devops middle+",
#     "Devops senior-",
#     "Devops senior",
#     "Devops senior+",
#     "Devops teamlead",
#     "Devops middle",
#     "Devops senior",
#     "Java middle",
#     "Java senior",
#     ".NET",
#     ".NET Senior",
#     "Security Specialist",
#     "Lead Programmer Researcher / AI",
#     "Programmer researcher / AI",
#     "Senior data analyst / AI",
#     "Middle data analyst / AI",
#     "Junior data analyst / AI",
#     "PM / AI",
#     "Teamlead / AI",
#     "Product analyst / AI",
#     "Solution Architect consultant / AI",
#     "Pr-manager junior",
#     "Pr-manager middle",
#     "Pr-manager senior",
#     "DevRel junior",
#     "DevRel middle",
#     "DevRel senior",
#     "Copywriter middle",
#     "Copywriter senior",
#     "Photographer/Videographer",
#     "Copywriter middle",
#     "Copywriter senior",
#     "Content Manager",
# ]

# #############################
# # 2. Чтение Excel (с формулами как значениями)
# #############################
# def read_excel_data_only(file, sheet_name=0):
#     file_data = file.read()
#     wb = openpyxl.load_workbook(BytesIO(file_data), data_only=True)
#     if isinstance(sheet_name, int):
#         sheet = wb.worksheets[sheet_name]
#     else:
#         sheet = wb[sheet_name]
#     data = list(sheet.values)
#     df = pd.DataFrame(data)
#     return df

# #############################
# # 3. Проверка грейдов
# #############################
# def check_grades(df, allowed_grades):
#     unknown_grades = set()
#     for col in df.columns:
#         if len(df) > 2:
#             third_row_value = df.iloc[2, col]
#             if pd.notna(third_row_value):
#                 text = str(third_row_value).strip().lower()
#                 if text in ["inside", "outside"]:
#                     grade_val = df.iloc[1, col]
#                     if pd.notna(grade_val):
#                         grade_str = str(grade_val).strip()
#                         if grade_str not in allowed_grades:
#                             unknown_grades.add(grade_str)
#     return unknown_grades

# #############################
# # 4. Поиск столбца "Total cost"
# #############################
# def find_total_cost_column_name(df):
#     for col in df.columns:
#         cell_value = df.iloc[1, col]
#         if pd.notna(cell_value) and str(cell_value).strip() == "Total cost":
#             return col
#     return None

# #############################
# # 5. Функция для обработки имени эпика
# #############################
# def process_function_name(epic_name):
#     # Заменяем пробелы на нижние подчёркивания
#     return "_".join(epic_name.split())

# #############################
# # 6. Жёстко берём столбцы F..Y как "оценочные"
# #############################
# def get_time_estimate_columns(df):
#     """
#     Возвращает список кортежей (grade_name, col_index) для столбцов F..Y (индексы 5..24),
#     где во 2-й строке (index=1) указано название грейда,
#     а в 3-й строке (index=2) написано 'Inside' или 'Outside'.
#     """
#     grade_cols = []
#     for col in range(5, 25):  # F..Y
#         if col < len(df.columns):
#             third_row_value = df.iloc[2, col]
#             if pd.notna(third_row_value):
#                 text = str(third_row_value).strip().lower()
#                 if text in ["inside", "outside"]:
#                     grade_val = df.iloc[1, col]
#                     if pd.notna(grade_val):
#                         grade_str = str(grade_val).strip()
#                         grade_cols.append((grade_str, col))
#     return grade_cols

# #############################
# # 7. Обработка "с эпиками"
# #############################
# def process_with_epics(df):
#     total_cost_col = find_total_cost_column_name(df)
#     grade_cols = get_time_estimate_columns(df)
#     start_row = 7  # Подберите под структуру

#     # Подготовка df_subset
#     df_subset = df.iloc[start_row:, [1, 2]].dropna(how='all').reset_index(drop=True)
#     df_subset.columns = ['Feature', 'Details']

#     # Основные списки
#     summary_list = []
#     custom_link_id_list = []
#     parent_link_id_list = []
#     issue_type_list = []
#     total_cost_list = []
#     function_name_list = []

#     # Списки для грейдов
#     grade_values = {grade_name: [] for (grade_name, _) in grade_cols}

#     current_custom_link_id = None
#     current_function_name = None

#     # -- Основной цикл --
#     for idx, row in df_subset.iterrows():
#         feature = row['Feature']
#         detail = row['Details']
#         original_row_index = idx + start_row

#         # Смотрим на Total cost
#         cost_value = None
#         if total_cost_col is not None and original_row_index < len(df):
#             cost_value = df.iloc[original_row_index, total_cost_col]

#         # Проверяем, является ли строка эпиком
#         if pd.notna(feature) and str(feature).strip() != "":
#             # Создаём строку для эпика
#             custom_id = str(random.randint(100000, 999999))
#             processed_fn = process_function_name(str(feature))

#             summary_list.append(feature)  # Summary = сам эпик
#             issue_type_list.append("Epic")
#             custom_link_id_list.append(custom_id)
#             parent_link_id_list.append(None)
#             total_cost_list.append(None)  # у эпика нет total cost
#             function_name_list.append(processed_fn)

#             current_custom_link_id = custom_id
#             current_function_name = processed_fn

#             # Времязатраты для эпика = None
#             for gname in grade_values:
#                 grade_values[gname].append(None)

#         # Проверяем, является ли строка ФТ
#         if pd.notna(detail) and str(detail).strip() != "":
#             # Создаём строку для ФТ
#             summary_list.append(detail)  # Summary = только столбец C
#             issue_type_list.append("ФТ")
#             custom_link_id_list.append(None)
#             parent_link_id_list.append(current_custom_link_id)
#             total_cost_list.append(cost_value if pd.notna(cost_value) else None)
#             function_name_list.append(current_function_name if current_function_name else None)

#             # Времязатраты для ФТ
#             for (grade_name, col_index) in grade_cols:
#                 if original_row_index < len(df):
#                     val = df.iloc[original_row_index, col_index]
#                     if pd.notna(val) and float(val) != 0.0:
#                         grade_values[grade_name].append(val)
#                     else:
#                         grade_values[grade_name].append(None)

#     # -- Формируем итоговый DataFrame --
#     result_df = pd.DataFrame({
#         'Summary': summary_list,
#         'Custom Link ID': custom_link_id_list,
#         'Parent Link ID': parent_link_id_list,
#         'Issue Type': issue_type_list,
#         'Total cost': total_cost_list,
#         'Function name': function_name_list
#     })

#     # -- Отладка: выводим длины списков в Streamlit --
#     st.write("### Отладочная информация (с эпиками)")
#     st.write(f"Итоговое количество строк в result_df: {len(result_df)}")
#     st.write(f"Количество строк в summary_list: {len(summary_list)}")
#     for grade_name, _ in grade_cols:
#         st.write(f"Грейд '{grade_name}': {len(grade_values[grade_name])} записей")

#     # -- Добавляем столбцы по грейдам --
#     for (grade_name, _) in grade_cols:
#         # Если выяснится, что длина не совпадает, вы это увидите в отладке
#         result_df[grade_name] = grade_values[grade_name]

#     return result_df

# #############################
# # 8. Обработка "без эпиков"
# #############################
# def process_without_epics(df):
#     total_cost_col = find_total_cost_column_name(df)
#     grade_cols = get_time_estimate_columns(df)
#     start_row = 7

#     df_subset = df.iloc[start_row:, [1, 2]].dropna(how='all').reset_index(drop=True)
#     df_subset.columns = ['Feature', 'Details']

#     summary_list = []
#     total_cost_list = []
#     function_name_list = []
#     issue_type_list = []

#     grade_values = {grade_name: [] for (grade_name, _) in grade_cols}

#     current_function_name = None

#     for idx, row in df_subset.iterrows():
#         feature = row['Feature']
#         detail = row['Details']
#         original_row_index = idx + start_row

#         cost_value = None
#         if total_cost_col is not None and original_row_index < len(df):
#             cost_value = df.iloc[original_row_index, total_cost_col]

#         # Если есть Feature (эпик), запоминаем, но не создаём отдельную строку
#         if pd.notna(feature) and str(feature).strip() != "":
#             current_function_name = process_function_name(str(feature))

#         # Если есть Details (ФТ)
#         if pd.notna(detail) and str(detail).strip() != "":
#             summary_list.append(detail)
#             total_cost_list.append(cost_value if pd.notna(cost_value) else None)
#             function_name_list.append(current_function_name if current_function_name else None)
#             issue_type_list.append("ФТ")

#             # Для каждого грейда
#             for (grade_name, col_index) in grade_cols:
#                 if original_row_index < len(df):
#                     val = df.iloc[original_row_index, col_index]
#                     if pd.notna(val) and float(val) != 0.0:
#                         grade_values[grade_name].append(val)
#                     else:
#                         grade_values[grade_name].append(None)
#         else:
#             # Если строка вообще не содержит ФТ (detail), 
#             # то не добавляем строку, следовательно, не добавляем в grade_values
#             # никаких значений.
#             pass

#     result_df = pd.DataFrame({
#         'Summary': summary_list,
#         'Issue Type': issue_type_list,
#         'Total cost': total_cost_list,
#         'Function name': function_name_list
#     })

#     # -- Отладка: выводим длины списков в Streamlit --
#     st.write("### Отладочная информация (без эпиков)")
#     st.write(f"Итоговое количество строк в result_df: {len(result_df)}")
#     st.write(f"Количество строк в summary_list: {len(summary_list)}")
#     for grade_name, _ in grade_cols:
#         st.write(f"Грейд '{grade_name}': {len(grade_values[grade_name])} записей")

#     # Добавляем столбцы по грейдам
#     for (grade_name, _) in grade_cols:
#         result_df[grade_name] = grade_values[grade_name]

#     return result_df

# #############################
# # 9. Основной поток (Streamlit)
# #############################
# def main():
#     st.title("Jira CSV Generator")

#     processing_option = st.radio(
#         "Выберите вариант обработки данных:",
#         ("Импортировать Функции как Epic's", "Не импортировать Эпики")
#     )

#     uploaded_file = st.file_uploader("Загрузите Excel файл", type=["xlsx"])
#     if uploaded_file:
#         st.success("Файл успешно загружен!")
#         df = read_excel_data_only(uploaded_file, sheet_name=0)

#         # Проверяем грейды
#         unknown_grades = check_grades(df, allowed_grades)
#         if unknown_grades:
#             st.warning(
#                 "Внимание! В смете присутствуют неизвестные грейды: " + ", ".join(unknown_grades)
#             )

#         if processing_option == "Импортировать Функции как Epic's":
#             result_df = process_with_epics(df)
#         else:
#             result_df = process_without_epics(df)

#         # Показываем DataFrame
#         st.dataframe(result_df)

#         # Скачивание CSV
#         csv = result_df.to_csv(index=False).encode('utf-8')
#         st.download_button(
#             label="Скачать CSV файл",
#             data=csv,
#             file_name='Jira-Import.csv',
#             mime='text/csv'
#         )

#     # Кнопка для скачивания конфиг-файла
#     config_file_path = "Конфиг v2.txt"
#     try:
#         with open(config_file_path, 'r') as config_file:
#             config_data = config_file.read()
#         st.download_button(
#             label="Скачать конфиг-файл для быстрого импорта",
#             data=config_data,
#             file_name='Jira-Import-Config.txt',
#             mime='text/plain'
#         )
#     except FileNotFoundError:
#         st.error(f"Файл {config_file_path} не найден. Убедитесь, что он загружен в репозиторий.")


# if __name__ == "__main__":
#     main()








# ВЕРСИЯ 0.3

# import streamlit as st
# import pandas as pd
# import random
# from io import BytesIO
# import openpyxl

# #############################
# # 1. Список допустимых грейдов (оставляем для проверки)
# #############################
# allowed_grades = [
#     "Продуктовая аналитика стажер",
#     "UX-аналитик junior",
#     "UX-аналитик middle -",
#     "UX-аналитик middle",
#     "UX-аналитик senior",
#     "Web/mobile-аналитик junior",
#     "Web/mobile-аналитик middle -",
#     "Web/mobile-аналитик middle",
#     "Web/mobile-аналитик middle +",
#     "Web/mobile-аналитик senior",
#     "Web/mobile-аналитик TL",
#     "BI/data-аналитик junior",
#     "BI/data-аналитик middle",
#     "BI/data-аналитик middle +",
#     "BI/data-аналитик senior",
#     "Product manager junior",
#     "Product manager middle",
#     "Product manager middle +",
#     "Product manager Senior",
#     "Web/mobile-analyst middle",
#     "Web/mobile-analyst senior",
#     "Дизайнер стажер",
#     "Дизайнер junior-",
#     "Дизайнер junior",
#     "Дизайнер junior+",
#     "Дизайнер middle-",
#     "Дизайнер middle",
#     "Дизайнер middle+",
#     "Дизайнер senior",
#     "Дизайнер senior+",
#     "Art Director",
#     "Web middle",
#     "Web senior",
#     "Системный аналитик Стажер",
#     "Системный аналитик junior-",
#     "Системный аналитик junior",
#     "Системный аналитик junior+",
#     "Системный аналитик middle-",
#     "Системный аналитик middle",
#     "Системный аналитик middle+",
#     "Системный аналитик senior",
#     "Системный аналитик senior+",
#     "Системный аналитик lead",
#     "Проектировщик стажер",
#     "Проектировщик junior",
#     "Проектировщик middle",
#     "Проектировщик middle+",
#     "Проектировщик senior",
#     "Проектировщик senior+",
#     "Проектировщик lead",
#     "System analyst middle",
#     "System analyst senior",
#     "SEO",
#     "tech.writer",
#     "UX writer",
#     "PM стажер",
#     "PM intern",
#     "PM junior",
#     "PM junior+",
#     "PM middle",
#     "PM middle+",
#     "PM senior",
#     "PM senior+",
#     "GH",
#     "PMO",
#     "Bitrix junior",
#     "Bitrix junior+",
#     "Bitrix middle-",
#     "Bitrix middle",
#     "Bitrix middle+",
#     "Bitrix senior-",
#     "Bitrix senior",
#     "Bitrix senior+",
#     "Bitrix teamlead -",
#     "Bitrix teamlead",
#     "Bitrix teamlead +",
#     "Bitrix teamlead grouphead -",
#     "Bitrix teamlead grouphead",
#     "Bitrix teamlead grouphead +",
#     "Bitrix middle",
#     "Bitrix senior",
#     "Разработка стажер",
#     "Framework junior",
#     "Framework junior+",
#     "Framework middle-",
#     "Framework middle",
#     "Framework middle+",
#     "Framework senior-",
#     "Framework senior",
#     "Framework senior+",
#     "Framework teamlead -",
#     "Framework teamlead",
#     "Framework teamlead +",
#     "Framework teamlead grouphead -",
#     "Framework teamlead grouphead",
#     "Framework teamlead grouphead +",
#     "Framework middle",
#     "Framework senior",
#     "QA junior-",
#     "QA junior",
#     "QA junior +",
#     "QA middle-",
#     "QA middle",
#     "QA middle+",
#     "QA senior -",
#     "QA senior",
#     "QA Teamlead",
#     "AQA junior-",
#     "AQA junior",
#     "AQA junior +",
#     "AQA middle-",
#     "AQA middle",
#     "AQA middle+",
#     "AQA senior -",
#     "AQA senior",
#     "AQA teamlead",
#     "QA middle",
#     "QA senior",
#     "QA AT Middle",
#     "QA AT Senior",
#     "Front-end junior",
#     "Front-end junior +",
#     "Front-end middle -",
#     "Front-end middle",
#     "Front-end middle +",
#     "Front-end senior -",
#     "Front-end senior",
#     "Front-end senior +",
#     "Front-end teamlead",
#     "NodeJS junior",
#     "NodeJS junior +",
#     "NodeJS middle -",
#     "NodeJS middle",
#     "NodeJS middle +",
#     "NodeJS senior -",
#     "NodeJS senior",
#     "NodeJS senior +",
#     "NodeJS teamlead",
#     "Front-end html/css middle -",
#     "Front-end html/css middle",
#     "Front-end html/css middle +",
#     "Front-end middle",
#     "Front-end senior",
#     "NodeJS middle",
#     "NodeJS senior",
#     "Front-end html/css middle",
#     "Mobile Dev Junior-",
#     "Mobile Dev Junior",
#     "Mobile Dev Junior+",
#     "Mobile Dev Middle-",
#     "Mobile Dev Middle",
#     "Mobile Dev Middle+",
#     "Mobile Dev Senior-",
#     "Mobile Dev Senior",
#     "Mobile Dev Senior+",
#     "Mobile Dev Teamlead",
#     "Mobile dev middle",
#     "Mobile dev senior",
#     "Python junior",
#     "Python middle-",
#     "Python middle",
#     "Python middle+",
#     "Python senior-",
#     "Python senior",
#     "Python senior+",
#     "Python teamlead",
#     "Golang junior",
#     "Golang middle-",
#     "Golang middle",
#     "Golang middle+",
#     "Golang senior-",
#     "Golang senior",
#     "Golang senior+",
#     "Golang teamlead",
#     "Python middle",
#     "Python senior",
#     "Golang middle",
#     "Golang senior",
#     "Devops junior",
#     "Devops middle-",
#     "Devops middle",
#     "Devops middle+",
#     "Devops senior-",
#     "Devops senior",
#     "Devops senior+",
#     "Devops teamlead",
#     "Devops middle",
#     "Devops senior",
#     "Java middle",
#     "Java senior",
#     ".NET",
#     ".NET Senior",
#     "Security Specialist",
#     "Lead Programmer Researcher / AI",
#     "Programmer researcher / AI",
#     "Senior data analyst / AI",
#     "Middle data analyst / AI",
#     "Junior data analyst / AI",
#     "PM / AI",
#     "Teamlead / AI",
#     "Product analyst / AI",
#     "Solution Architect consultant / AI",
#     "Pr-manager junior",
#     "Pr-manager middle",
#     "Pr-manager senior",
#     "DevRel junior",
#     "DevRel middle",
#     "DevRel senior",
#     "Copywriter middle",
#     "Copywriter senior",
#     "Photographer/Videographer",
#     "Copywriter middle",
#     "Copywriter senior",
#     "Content Manager",
# ]

# #############################
# # 2. Чтение Excel (с формулами как значениями)
# #############################
# def read_excel_data_only(file, sheet_name=0):
#     file_data = file.read()
#     wb = openpyxl.load_workbook(BytesIO(file_data), data_only=True)
#     if isinstance(sheet_name, int):
#         sheet = wb.worksheets[sheet_name]
#     else:
#         sheet = wb[sheet_name]
#     data = list(sheet.values)
#     df = pd.DataFrame(data)
#     return df

# #############################
# # 3. Функция для проверки грейдов
# #############################
# def check_grades(df, allowed_grades):
#     unknown_grades = set()
#     for col in df.columns:
#         if len(df) > 2:
#             third_row_value = df.iloc[2, col]
#             if pd.notna(third_row_value):
#                 text = str(third_row_value).strip().lower()
#                 if text in ["inside", "outside"]:
#                     grade_val = df.iloc[1, col]
#                     if pd.notna(grade_val):
#                         grade_str = str(grade_val).strip()
#                         if grade_str not in allowed_grades:
#                             unknown_grades.add(grade_str)
#     return unknown_grades

# #############################
# # 4. Остальные функции
# #############################
# def find_total_cost_column_name(df):
#     for col in df.columns:
#         cell_value = df.iloc[1, col]
#         if pd.notna(cell_value) and str(cell_value).strip() == "Total cost":
#             return col
#     return None

# def process_function_name(epic_name):
#     # Заменяем пробелы на нижние подчёркивания
#     return "_".join(epic_name.split())

# #############################
# # 5. Обработка "с эпиками"
# #############################
# def process_with_epics(df):
#     total_cost_col = find_total_cost_column_name(df)
#     start_row = 7  # Подберите под вашу структуру

#     # Берём только столбцы B и C
#     df_subset = df.iloc[start_row:, [1, 2]].dropna(how='all').reset_index(drop=True)
#     df_subset.columns = ['Feature', 'Details']

#     summary_list = []
#     custom_link_id_list = []
#     parent_link_id_list = []
#     issue_type_list = []
#     total_cost_list = []
#     function_name_list = []  # новый столбец для Function name

#     current_custom_link_id = None
#     current_function_name = None

#     for idx, row in df_subset.iterrows():
#         feature = row['Feature']
#         detail = row['Details']
#         original_row_index = idx + start_row

#         cost_value = None
#         if total_cost_col is not None and original_row_index < len(df):
#             cost_value = df.iloc[original_row_index, total_cost_col]

#         # Если строка с Эпиком (Feature заполнен)
#         if pd.notna(feature):
#             custom_id = str(random.randint(100000, 999999))
#             processed_fn = process_function_name(str(feature))
#             summary_list.append(feature)  # для эпика оставляем исходное название
#             issue_type_list.append("Epic")
#             custom_link_id_list.append(custom_id)
#             parent_link_id_list.append(None)
#             total_cost_list.append(None)
#             function_name_list.append(processed_fn)
#             current_custom_link_id = custom_id
#             current_function_name = processed_fn

#         # Если строка с ФТ (Details заполнен)
#         if pd.notna(detail):
#             summary_list.append(detail)  # теперь только значение из столбца C
#             issue_type_list.append("ФТ")
#             custom_link_id_list.append(None)
#             parent_link_id_list.append(current_custom_link_id)
#             total_cost_list.append(cost_value if pd.notna(cost_value) else None)
#             # Для ФТ берём последнее известное значение Function name
#             function_name_list.append(current_function_name if current_function_name is not None else None)

#     result_df = pd.DataFrame({
#         'Summary': summary_list,
#         'Custom Link ID': custom_link_id_list,
#         'Parent Link ID': parent_link_id_list,
#         'Issue Type': issue_type_list,
#         'Total cost': total_cost_list,
#         'Function name': function_name_list
#     })

#     return result_df

# #############################
# # 6. Обработка "без эпиков"
# #############################
# def process_without_epics(df):
#     total_cost_col = find_total_cost_column_name(df)
#     start_row = 7

#     df_subset = df.iloc[start_row:, [1, 2]].dropna(how='all').reset_index(drop=True)
#     df_subset.columns = ['Feature', 'Details']

#     summary_list = []
#     total_cost_list = []
#     function_name_list = []
#     current_function_name = None

#     for idx, row in df_subset.iterrows():
#         original_row_index = idx + start_row
#         cost_value = None
#         if total_cost_col is not None and original_row_index < len(df):
#             cost_value = df.iloc[original_row_index, total_cost_col]

#         if pd.notna(row['Feature']):
#             current_function_name = process_function_name(str(row['Feature']))

#         if pd.notna(row['Details']):
#             summary_list.append(row['Details'])
#             total_cost_list.append(cost_value if pd.notna(cost_value) else None)
#             function_name_list.append(current_function_name if current_function_name is not None else None)

#     result_df = pd.DataFrame({
#         'Summary': summary_list,
#         'Issue Type': ['ФТ'] * len(summary_list),
#         'Total cost': total_cost_list,
#         'Function name': function_name_list
#     })

#     return result_df

# #############################
# # 7. Основной поток (Streamlit)
# #############################
# def main():
#     st.title("Jira CSV Generator")

#     processing_option = st.radio(
#         "Выберите вариант обработки данных:",
#         ("Импортировать Функции как Epic's", "Не импортировать Эпики")
#     )

#     uploaded_file = st.file_uploader("Загрузите Excel файл", type=["xlsx"])
#     if uploaded_file:
#         st.success("Файл успешно загружен!")
#         df = read_excel_data_only(uploaded_file, sheet_name=0)

#         # Проверяем грейды (для предупреждения, если найдены неизвестные)
#         unknown_grades = check_grades(df, allowed_grades)
#         if unknown_grades:
#             st.warning(
#                 "Внимание! В смете присутствуют неизвестные грейды: " + ", ".join(unknown_grades)
#             )

#         if processing_option == "Импортировать Функции как Epic's":
#             result_df = process_with_epics(df)
#         else:
#             result_df = process_without_epics(df)

#         st.dataframe(result_df)
#         csv = result_df.to_csv(index=False).encode('utf-8')
#         st.download_button(
#             label="Скачать CSV файл",
#             data=csv,
#             file_name='Jira-Import.csv',
#             mime='text/csv'
#         )

#     config_file_path = "Конфиг v2.txt"
#     try:
#         with open(config_file_path, 'r') as config_file:
#             config_data = config_file.read()
#         st.download_button(
#             label="Скачать конфиг-файл для быстрого импорта",
#             data=config_data,
#             file_name='Jira-Import-Config.txt',
#             mime='text/plain'
#         )
#     except FileNotFoundError:
#         st.error(f"Файл {config_file_path} не найден. Убедитесь, что он загружен в репозиторий.")

# if __name__ == "__main__":
#     main()









# ВЕРСИЯ 0.2

# import streamlit as st
# import pandas as pd
# import random
# from io import BytesIO
# import openpyxl

# #############################
# # 1. Список допустимых грейдов
# #############################
# allowed_grades = [
#     "Продуктовая аналитика стажер",
#     "UX-аналитик junior",
#     "UX-аналитик middle -",
#     "UX-аналитик middle",
#     "UX-аналитик senior",
#     "Web/mobile-аналитик junior",
#     "Web/mobile-аналитик middle -",
#     "Web/mobile-аналитик middle",
#     "Web/mobile-аналитик middle +",
#     "Web/mobile-аналитик senior",
#     "Web/mobile-аналитик TL",
#     "BI/data-аналитик junior",
#     "BI/data-аналитик middle",
#     "BI/data-аналитик middle +",
#     "BI/data-аналитик senior",
#     "Product manager junior",
#     "Product manager middle",
#     "Product manager middle +",
#     "Product manager Senior",
#     "Web/mobile-analyst middle",
#     "Web/mobile-analyst senior",
#     "Дизайнер стажер",
#     "Дизайнер junior-",
#     "Дизайнер junior",
#     "Дизайнер junior+",
#     "Дизайнер middle-",
#     "Дизайнер middle",
#     "Дизайнер middle+",
#     "Дизайнер senior",
#     "Дизайнер senior+",
#     "Art Director",
#     "Web middle",
#     "Web senior",
#     "Системный аналитик Стажер",
#     "Системный аналитик junior-",
#     "Системный аналитик junior",
#     "Системный аналитик junior+",
#     "Системный аналитик middle-",
#     "Системный аналитик middle",
#     "Системный аналитик middle+",
#     "Системный аналитик senior",
#     "Системный аналитик senior+",
#     "Системный аналитик lead",
#     "Проектировщик стажер",
#     "Проектировщик junior",
#     "Проектировщик middle",
#     "Проектировщик middle+",
#     "Проектировщик senior",
#     "Проектировщик senior+",
#     "Проектировщик lead",
#     "System analyst middle",
#     "System analyst senior",
#     "SEO",
#     "tech.writer",
#     "UX writer",
#     "PM стажер",
#     "PM intern",
#     "PM junior",
#     "PM junior+",
#     "PM middle",
#     "PM middle+",
#     "PM senior",
#     "PM senior+",
#     "GH",
#     "PMO",
#     "Bitrix junior",
#     "Bitrix junior+",
#     "Bitrix middle-",
#     "Bitrix middle",
#     "Bitrix middle+",
#     "Bitrix senior-",
#     "Bitrix senior",
#     "Bitrix senior+",
#     "Bitrix teamlead -",
#     "Bitrix teamlead",
#     "Bitrix teamlead +",
#     "Bitrix teamlead grouphead -",
#     "Bitrix teamlead grouphead",
#     "Bitrix teamlead grouphead +",
#     "Bitrix middle",
#     "Bitrix senior",
#     "Разработка стажер",
#     "Framework junior",
#     "Framework junior+",
#     "Framework middle-",
#     "Framework middle",
#     "Framework middle+",
#     "Framework senior-",
#     "Framework senior",
#     "Framework senior+",
#     "Framework teamlead -",
#     "Framework teamlead",
#     "Framework teamlead +",
#     "Framework teamlead grouphead -",
#     "Framework teamlead grouphead",
#     "Framework teamlead grouphead +",
#     "Framework middle",
#     "Framework senior",
#     "QA junior-",
#     "QA junior",
#     "QA junior +",
#     "QA middle-",
#     "QA middle",
#     "QA middle+",
#     "QA senior -",
#     "QA senior",
#     "QA Teamlead",
#     "AQA junior-",
#     "AQA junior",
#     "AQA junior +",
#     "AQA middle-",
#     "AQA middle",
#     "AQA middle+",
#     "AQA senior -",
#     "AQA senior",
#     "AQA teamlead",
#     "QA middle",
#     "QA senior",
#     "QA AT Middle",
#     "QA AT Senior",
#     "Front-end junior",
#     "Front-end junior +",
#     "Front-end middle -",
#     "Front-end middle",
#     "Front-end middle +",
#     "Front-end senior -",
#     "Front-end senior",
#     "Front-end senior +",
#     "Front-end teamlead",
#     "NodeJS junior",
#     "NodeJS junior +",
#     "NodeJS middle -",
#     "NodeJS middle",
#     "NodeJS middle +",
#     "NodeJS senior -",
#     "NodeJS senior",
#     "NodeJS senior +",
#     "NodeJS teamlead",
#     "Front-end html/css middle -",
#     "Front-end html/css middle",
#     "Front-end html/css middle +",
#     "Front-end middle",
#     "Front-end senior",
#     "NodeJS middle",
#     "NodeJS senior",
#     "Front-end html/css middle",
#     "Mobile Dev Junior-",
#     "Mobile Dev Junior",
#     "Mobile Dev Junior+",
#     "Mobile Dev Middle-",
#     "Mobile Dev Middle",
#     "Mobile Dev Middle+",
#     "Mobile Dev Senior-",
#     "Mobile Dev Senior",
#     "Mobile Dev Senior+",
#     "Mobile Dev Teamlead",
#     "Mobile dev middle",
#     "Mobile dev senior",
#     "Python junior",
#     "Python middle-",
#     "Python middle",
#     "Python middle+",
#     "Python senior-",
#     "Python senior",
#     "Python senior+",
#     "Python teamlead",
#     "Golang junior",
#     "Golang middle-",
#     "Golang middle",
#     "Golang middle+",
#     "Golang senior-",
#     "Golang senior",
#     "Golang senior+",
#     "Golang teamlead",
#     "Python middle",
#     "Python senior",
#     "Golang middle",
#     "Golang senior",
#     "Devops junior",
#     "Devops middle-",
#     "Devops middle",
#     "Devops middle+",
#     "Devops senior-",
#     "Devops senior",
#     "Devops senior+",
#     "Devops teamlead",
#     "Devops middle",
#     "Devops senior",
#     "Java middle",
#     "Java senior",
#     ".NET",
#     ".NET Senior",
#     "Security Specialist",
#     "Lead Programmer Researcher / AI",
#     "Programmer researcher / AI",
#     "Senior data analyst / AI",
#     "Middle data analyst / AI",
#     "Junior data analyst / AI",
#     "PM / AI",
#     "Teamlead / AI",
#     "Product analyst / AI",
#     "Solution Architect consultant / AI",
#     "Pr-manager junior",
#     "Pr-manager middle",
#     "Pr-manager senior",
#     "DevRel junior",
#     "DevRel middle",
#     "DevRel senior",
#     "Copywriter middle",
#     "Copywriter senior",
#     "Photographer/Videographer",
#     "Copywriter middle",
#     "Copywriter senior",
#     "Content Manager",
# ]

# #############################
# # 2. Чтение Excel (с формулами как значениями)
# #############################
# def read_excel_data_only(file, sheet_name=0):
#     file_data = file.read()
#     wb = openpyxl.load_workbook(BytesIO(file_data), data_only=True)
#     if isinstance(sheet_name, int):
#         sheet = wb.worksheets[sheet_name]
#     else:
#         sheet = wb[sheet_name]

#     data = list(sheet.values)
#     df = pd.DataFrame(data)
#     return df

# #############################
# # 3. Функция для проверки грейдов
# #############################
# def check_grades(df, allowed_grades):
#     """
#     Проверяем в 3-й строке (индекс 2) ячейки на значения "Inside" или "Outside".
#     Если находим такие ячейки, значит во 2-й строке (индекс 1) лежит грейд.
#     Сравниваем грейд со списком allowed_grades.
#     Если находим незнакомый грейд, собираем в множество unknown_grades.
#     Возвращаем множество unknown_grades (или пустое множество).
#     """
#     unknown_grades = set()

#     # Перебираем все столбцы df
#     for col in df.columns:
#         # Смотрим, что написано в 3-й строке (индекс 2)
#         if len(df) > 2:  # чтобы не выйти за границы, если таблица меньше 3 строк
#             third_row_value = df.iloc[2, col]
#             if pd.notna(third_row_value):
#                 text = str(third_row_value).strip().lower()
#                 # Если "Inside" или "Outside", значит в строке выше (index=1) лежит грейд
#                 if text in ["inside", "outside"]:
#                     # Читаем грейд
#                     grade_val = df.iloc[1, col]
#                     if pd.notna(grade_val):
#                         grade_str = str(grade_val).strip()
#                         # Сравниваем с списком
#                         if grade_str not in allowed_grades:
#                             unknown_grades.add(grade_str)

#     return unknown_grades

# #############################
# # 4. Остальные функции
# #############################
# def find_total_cost_column_name(df):
#     for col in df.columns:
#         cell_value = df.iloc[1, col]
#         if pd.notna(cell_value) and str(cell_value).strip() == "Total cost":
#             return col
#     return None

# def process_with_epics(df):
#     total_cost_col = find_total_cost_column_name(df)
#     start_row = 7  # TODO

#     df_subset = df.iloc[start_row:, [1, 2]].dropna(how='all').reset_index(drop=True)
#     df_subset.columns = ['Feature', 'Details']

#     summary_list = []
#     custom_link_id_list = []
#     parent_link_id_list = []
#     issue_type_list = []
#     total_cost_list = []

#     current_custom_link_id = None
#     current_epic_name = None

#     for idx, row in df_subset.iterrows():
#         feature = row['Feature']
#         detail = row['Details']
#         original_row_index = idx + start_row

#         cost_value = None
#         if total_cost_col is not None and original_row_index < len(df):
#             cost_value = df.iloc[original_row_index, total_cost_col]

#         if pd.notna(feature):
#             summary_list.append(feature)
#             issue_type_list.append("Epic")
#             custom_id = str(random.randint(100000, 999999))
#             custom_link_id_list.append(custom_id)
#             parent_link_id_list.append(None)
#             total_cost_list.append(None)
#             current_custom_link_id = custom_id
#             current_epic_name = feature

#         if pd.notna(detail):
#             if current_epic_name:
#                 ft_summary = f"[{current_epic_name}] {detail}"
#             else:
#                 ft_summary = detail

#             summary_list.append(ft_summary)
#             issue_type_list.append("ФТ")
#             custom_link_id_list.append(None)
#             parent_link_id_list.append(current_custom_link_id)
#             total_cost_list.append(cost_value if pd.notna(cost_value) else None)

#     return pd.DataFrame({
#         'Summary': summary_list,
#         'Custom Link ID': custom_link_id_list,
#         'Parent Link ID': parent_link_id_list,
#         'Issue Type': issue_type_list,
#         'Total cost': total_cost_list
#     })

# def process_without_epics(df):
#     total_cost_col = find_total_cost_column_name(df)
#     start_row = 7

#     df_subset = df.iloc[start_row:, [1, 2]].dropna(how='all').reset_index(drop=True)
#     df_subset.columns = ['Feature', 'Details']

#     summary_list = []
#     total_cost_list = []
#     current_epic_name = None

#     for idx, row in df_subset.iterrows():
#         feature = row['Feature']
#         detail = row['Details']
#         original_row_index = idx + start_row

#         cost_value = None
#         if total_cost_col is not None and original_row_index < len(df):
#             cost_value = df.iloc[original_row_index, total_cost_col]

#         if pd.notna(feature):
#             current_epic_name = feature

#         if pd.notna(detail):
#             if current_epic_name:
#                 summary = f"[{current_epic_name}] {detail}"
#             else:
#                 summary = detail

#             summary_list.append(summary)
#             total_cost_list.append(cost_value if pd.notna(cost_value) else None)

#     return pd.DataFrame({
#         'Summary': summary_list,
#         'Issue Type': ['ФТ'] * len(summary_list),
#         'Total cost': total_cost_list
#     })

# #############################
# # 5. Основной поток (Streamlit)
# #############################
# def main():
#     st.title("Jira CSV Generator")

#     processing_option = st.radio(
#         "Выберите вариант обработки данных:",
#         ("Импортировать Функции как Epic's", "Не импортировать Эпики")
#     )

#     uploaded_file = st.file_uploader("Загрузите Excel файл", type=["xlsx"])
#     if uploaded_file:
#         st.success("Файл успешно загружен!")
#         df = read_excel_data_only(uploaded_file, sheet_name=0)

#         # Сначала проверяем грейды
#         unknown_grades = check_grades(df, allowed_grades)
#         if unknown_grades:
#             st.warning(
#                 "Внимание! В смете присутствуют неизвестные грейды: "
#                 + ", ".join(unknown_grades)
#             )

#         # Затем продолжаем обычную обработку
#         if processing_option == "Импортировать Функции как Epic's":
#             result_df = process_with_epics(df)
#         else:
#             result_df = process_without_epics(df)

#         st.dataframe(result_df)
#         csv = result_df.to_csv(index=False).encode('utf-8')
#         st.download_button(
#             label="Скачать CSV файл",
#             data=csv,
#             file_name='Jira-Import.csv',
#             mime='text/csv'
#         )

#     config_file_path = "Конфиг v2.txt"
#     try:
#         with open(config_file_path, 'r') as config_file:
#             config_data = config_file.read()
#         st.download_button(
#             label="Скачать конфиг-файл для быстрого импорта",
#             data=config_data,
#             file_name='Jira-Import-Config.txt',
#             mime='text/plain'
#         )
#     except FileNotFoundError:
#         st.error(f"Файл {config_file_path} не найден. Убедитесь, что он загружен в репозиторий.")

# if __name__ == "__main__":
#     main()
